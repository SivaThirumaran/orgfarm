import openpyxl

def getRowCount(file,Sheetname):
    workbook = openpyxl.load_workbook(file)
    Sheet = workbook[Sheetname]
    return(Sheet.max_row)

def getColumnCount(file,Sheetname):
    workbook = openpyxl.load_workbook(file)
    Sheet = workbook[Sheetname]
    return (Sheet.max_column)

def readData(file,Sheetname,romnum,columnno):
    workbook = openpyxl.load_workbook(file)
    Sheet = workbook[Sheetname]
    return Sheet.cell(row=romnum, column=columnno).value

def writeData(file,Sheetname,romnum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    Sheet = workbook[Sheetname]
    Sheet.cell(row=romnum, column=columnno).value = data
    workbook.save(file)